"""
ETABS Seismic Optimization Tool - Phase 1
Step 8: Logging Results
========================
Tracks every evaluate_design() call in a pandas DataFrame.
Auto-saves to Excel every N runs so crashes don't lose data.
"""

import os
import json
import time
from datetime import datetime
from typing import Dict, List, Optional, Any

import pandas as pd


# ─── Column definitions ───────────────────────────────────────────────────────
#
# Grouped into 4 sections:
#   A. Run metadata
#   B. Input (section sizes applied)
#   C. Structural results
#   D. Compliance flags

LOG_COLUMNS = [
    # A. Metadata
    "run_id",           # sequential integer
    "timestamp",        # datetime string
    "run_time_sec",     # wall time for this evaluation
    "error_msg",        # "" = clean run

    # B. Section inputs — expand as needed for your design groups
    # These are stored as a JSON string of the full design_config dict
    "design_config_json",

    # Also store key section sizes as individual columns for easy filtering:
    # (Add columns matching your actual design group names)
    "grp_story1_col",   # e.g. "C30x30"
    "grp_story2_col",
    "grp_story3_col",
    "grp_story1_beam",
    "grp_story2_beam",
    "grp_story3_beam",

    # C. Structural results
    "T1_sec",
    "T2_sec",
    "SumUX_pct",
    "SumUY_pct",
    "max_drift_X",
    "max_drift_Y",
    "Vx_kip",
    "Vy_kip",
    "max_col_pmm_dcr",
    "n_beam_fail",
    "n_col_fail",

    # Per-story drifts — stored as JSON strings
    "story_drifts_X_json",
    "story_drifts_Y_json",

    # D. Compliance flags
    "drift_pass_X",
    "drift_pass_Y",
    "analysis_ok",
    "overall_pass",
]


class RunLogger:
    """
    Manages the results log DataFrame and auto-saves to Excel.

    Usage:
        logger = RunLogger(output_dir="results", save_every=10)
        ...
        logger.log(result_dict)
        ...
        logger.save()   # final save
    """

    def __init__(
        self,
        output_dir: str = "results",
        project_name: str = "seismic_opt",
        save_every: int = 10,
        resume_file: Optional[str] = None
    ):
        self.output_dir   = output_dir
        self.project_name = project_name
        self.save_every   = save_every
        self._run_counter = 0
        self._records: List[dict] = []

        os.makedirs(output_dir, exist_ok=True)

        # Resume from existing file if provided
        if resume_file and os.path.isfile(resume_file):
            existing = pd.read_excel(resume_file)
            self._records = existing.to_dict('records')
            self._run_counter = len(self._records)
            print(f"  ✓ Resumed {self._run_counter} existing runs from {resume_file}")

        self._session_start = datetime.now().strftime("%Y%m%d_%H%M%S")

    @property
    def excel_path(self) -> str:
        return os.path.join(
            self.output_dir,
            f"{self.project_name}_{self._session_start}.xlsx"
        )

    def log(self, result: Dict[str, Any]) -> int:
        """
        Add one evaluate_design() result to the log.
        Triggers an auto-save every `save_every` runs.
        Returns the run_id assigned.
        """
        self._run_counter += 1
        run_id = self._run_counter

        dc = result.get('design_config', {}) or {}

        record = {
            # Metadata
            "run_id":          run_id,
            "timestamp":       datetime.now().isoformat(timespec='seconds'),
            "run_time_sec":    result.get('run_time_sec', 0),
            "error_msg":       result.get('error_msg', ''),

            # Design config
            "design_config_json": json.dumps(dc),

            # Expand key groups into individual columns
            # Update these keys to match YOUR group names
            "grp_story1_col":   dc.get("Story1_Columns", ""),
            "grp_story2_col":   dc.get("Story2_Columns", ""),
            "grp_story3_col":   dc.get("Story3_Columns", ""),
            "grp_story1_beam":  dc.get("Story1_Beams", ""),
            "grp_story2_beam":  dc.get("Story2_Beams", ""),
            "grp_story3_beam":  dc.get("Story3_Beams", ""),

            # Structural results
            "T1_sec":           result.get('T1_sec', 0),
            "T2_sec":           result.get('T2_sec', 0),
            "SumUX_pct":        result.get('SumUX_pct', 0),
            "SumUY_pct":        result.get('SumUY_pct', 0),
            "max_drift_X":      result.get('max_drift_X', 999),
            "max_drift_Y":      result.get('max_drift_Y', 999),
            "Vx_kip":           result.get('Vx_kip', 0),
            "Vy_kip":           result.get('Vy_kip', 0),
            "max_col_pmm_dcr":  result.get('max_col_pmm_dcr', 999),
            "n_beam_fail":      result.get('n_beam_fail', 999),
            "n_col_fail":       result.get('n_col_fail', 999),

            # Per-story drifts as JSON
            "story_drifts_X_json": json.dumps(result.get('story_drifts_X', {})),
            "story_drifts_Y_json": json.dumps(result.get('story_drifts_Y', {})),

            # Compliance
            "drift_pass_X":    result.get('drift_pass_X', False),
            "drift_pass_Y":    result.get('drift_pass_Y', False),
            "analysis_ok":     result.get('analysis_ok', False),
            "overall_pass":    result.get('overall_pass', False),
        }

        self._records.append(record)

        # Auto-save
        if run_id % self.save_every == 0:
            self.save(verbose=True)

        return run_id

    def to_dataframe(self) -> pd.DataFrame:
        """Return all logged runs as a DataFrame."""
        if not self._records:
            return pd.DataFrame(columns=LOG_COLUMNS)
        return pd.DataFrame(self._records)

    def save(self, path: Optional[str] = None, verbose: bool = True) -> str:
        """
        Save the log to Excel.
        Uses openpyxl engine; applies basic formatting.
        Returns the file path written.
        """
        path = path or self.excel_path
        df = self.to_dataframe()

        try:
            with pd.ExcelWriter(path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Results', index=False)

                # Auto-fit column widths
                ws = writer.sheets['Results']
                for col in ws.columns:
                    max_len = max(
                        len(str(cell.value)) if cell.value else 0
                        for cell in col
                    )
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

                # Freeze top row
                ws.freeze_panes = 'A2'

                # Highlight failing runs in light red
                from openpyxl.styles import PatternFill
                red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC',
                                       fill_type='solid')
                pass_col_idx = df.columns.get_loc('overall_pass') + 1
                for row_idx, row_val in enumerate(df['overall_pass'], start=2):
                    if not row_val:
                        for cell in ws[row_idx]:
                            cell.fill = red_fill

        except Exception as e:
            print(f"  ⚠ Save failed: {e}")
            # Fallback: save as CSV
            csv_path = path.replace('.xlsx', '.csv')
            df.to_csv(csv_path, index=False)
            print(f"  → Saved as CSV: {csv_path}")
            return csv_path

        if verbose:
            print(f"  ✓ Saved {len(df)} runs → {path}")
        return path

    def save_best(self, n: int = 10, path: Optional[str] = None) -> pd.DataFrame:
        """
        Save the top N passing designs sorted by:
          1. overall_pass == True
          2. max(drift_X, drift_Y) ascending
          3. max_col_pmm_dcr ascending
        Returns the filtered/sorted DataFrame.
        """
        df = self.to_dataframe()
        passing = df[df['overall_pass'] == True].copy()

        if passing.empty:
            print("  ⚠ No passing designs to rank yet.")
            return passing

        passing['max_drift'] = passing[['max_drift_X', 'max_drift_Y']].max(axis=1)
        best = passing.sort_values(
            ['max_drift', 'max_col_pmm_dcr']
        ).head(n)

        if path:
            best.to_excel(path, index=False)
            print(f"  ✓ Saved top {len(best)} designs → {path}")

        return best

    def print_summary(self) -> None:
        """Print a quick console summary of all runs so far."""
        df = self.to_dataframe()
        if df.empty:
            print("  No runs logged yet.")
            return

        n_total  = len(df)
        n_pass   = df['overall_pass'].sum()
        n_fail   = n_total - n_pass
        avg_time = df['run_time_sec'].mean()

        print(f"\n{'─'*50}")
        print(f"  Runs: {n_total}  |  Pass: {n_pass}  |  Fail: {n_fail}")
        print(f"  Avg run time: {avg_time:.1f}s")
        if n_pass > 0:
            best = df[df['overall_pass']].sort_values('max_drift_X').iloc[0]
            print(f"  Best so far: run #{best['run_id']}")
            print(f"    drift X={best['max_drift_X']:.4f}  Y={best['max_drift_Y']:.4f}")
            print(f"    T1={best['T1_sec']:.3f}s  col_PMM={best['max_col_pmm_dcr']:.3f}")
        print(f"{'─'*50}\n")


# ─── Standalone utility: load a saved log ────────────────────────────────────

def load_log(path: str) -> pd.DataFrame:
    """Load a previously saved results Excel file."""
    df = pd.read_excel(path)
    print(f"Loaded {len(df)} runs from {path}")
    return df


# ─── Test ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    # Simulate logging 15 fake results
    logger = RunLogger(output_dir="results", project_name="test_run", save_every=5)

    import random
    for i in range(15):
        fake_result = {
            'design_config': {
                'Story1_Columns': random.choice(['C24x24', 'C30x30']),
                'Story2_Columns': random.choice(['C24x24', 'C30x30']),
                'Story1_Beams':   random.choice(['B18x24', 'B21x27']),
                'Story2_Beams':   random.choice(['B18x24', 'B21x27']),
            },
            'analysis_ok':      True,
            'max_drift_X':      random.uniform(0.008, 0.025),
            'max_drift_Y':      random.uniform(0.008, 0.025),
            'story_drifts_X':   {'Story1': 0.012, 'Story2': 0.018},
            'story_drifts_Y':   {'Story1': 0.010, 'Story2': 0.015},
            'Vx_kip':           random.uniform(200, 600),
            'Vy_kip':           random.uniform(200, 600),
            'T1_sec':           random.uniform(0.5, 2.0),
            'T2_sec':           random.uniform(0.3, 1.5),
            'SumUX_pct':        random.uniform(80, 95),
            'SumUY_pct':        random.uniform(80, 95),
            'max_col_pmm_dcr':  random.uniform(0.4, 1.2),
            'n_beam_fail':      0,
            'n_col_fail':       0,
            'drift_pass_X':     True,
            'drift_pass_Y':     True,
            'overall_pass':     random.choice([True, True, False]),
            'run_time_sec':     random.uniform(20, 120),
            'error_msg':        '',
        }
        run_id = logger.log(fake_result)
        print(f"  Logged run #{run_id}  pass={fake_result['overall_pass']}")

    logger.save()
    logger.print_summary()

    best = logger.save_best(n=5, path="results/best_designs.xlsx")
    print("\nTop 5 designs:")
    print(best[['run_id', 'max_drift_X', 'max_drift_Y', 'max_col_pmm_dcr']].to_string(index=False))
